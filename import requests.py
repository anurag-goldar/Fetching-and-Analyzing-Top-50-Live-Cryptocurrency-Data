import requests
import pandas as pd
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Step 1: Fetch Live Data
def fetch_top_50_cryptos():
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": 50,
        "page": 1,
        "sparkline": False
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        data = response.json()
        return data
    else:
        print("Failed to fetch data")
        return None

# Process fetched data into a DataFrame
def process_data(data):
    cryptos = []
    for crypto in data:
        cryptos.append({
            "Name": crypto["name"],
            "Symbol": crypto["symbol"],
            "Price (USD)": crypto["current_price"],
            "Market Cap": crypto["market_cap"],
            "24h Volume": crypto["total_volume"],
            "24h Price Change (%)": crypto["price_change_percentage_24h"]
        })
    return pd.DataFrame(cryptos)

# Save data and analysis to Excel with formatting
def save_to_excel(df, analysis_results, top_5_by_market_cap, filename="crypto_data.xlsx"):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        # Save the main data to the first sheet
        df.to_excel(writer, sheet_name="Cryptocurrency Data", index=False)
        
        # Save the analysis results to the second sheet
        analysis_results.to_excel(writer, sheet_name="Analysis Results", index=False, startrow=1)
        
        # Save the Top 5 Cryptocurrencies by Market Cap as a separate table
        top_5_by_market_cap.to_excel(writer, sheet_name="Analysis Results", index=False, startrow=len(analysis_results) + 4)
        
        # Access the workbook and worksheet for formatting
        workbook = writer.book
        worksheet = writer.sheets["Analysis Results"]
        
        # Formatting the Analysis Results sheet
        header_font = Font(bold=True, size=12)
        alignment = Alignment(horizontal="left", vertical="center")
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 60
        
        # Format headers for analysis results
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.alignment = alignment
        
        # Format data rows for analysis results
        for row in worksheet.iter_rows(min_row=2, max_row=len(analysis_results) + 1, max_col=2):
            for cell in row:
                cell.alignment = alignment
                cell.border = border
        
        # Add a title for the Top 5 Cryptocurrencies table
        worksheet.cell(row=len(analysis_results) + 3, column=1, value="Top 5 Cryptocurrencies by Market Cap").font = header_font
        
        # Format headers for the Top 5 Cryptocurrencies table
        for cell in worksheet[worksheet.max_row - len(top_5_by_market_cap):worksheet.max_row]:
            for c in cell:
                c.font = header_font
                c.alignment = alignment
                c.border = border

# Step 2: Data Analysis
def analyze_data(df):
    # Top 5 by Market Cap
    top_5_by_market_cap = df.nlargest(5, "Market Cap")[["Name", "Market Cap"]]
    
    # Average Price
    average_price = df["Price (USD)"].mean()
    
    # Highest and Lowest 24h Price Change
    highest_price_change = df.nlargest(1, "24h Price Change (%)")[["Name", "24h Price Change (%)"]]
    lowest_price_change = df.nsmallest(1, "24h Price Change (%)")[["Name", "24h Price Change (%)"]]
    
    # Combine analysis results into a DataFrame
    analysis_results = pd.DataFrame({
        "Metric": [
            "Average Price of Top 50 Cryptocurrencies",
            "Highest 24h Price Change",
            "Lowest 24h Price Change"
        ],
        "Value": [
            f"${average_price:.2f}",
            highest_price_change.to_string(index=False),
            lowest_price_change.to_string(index=False)
        ]
    })
    
    return analysis_results, top_5_by_market_cap

# Main function to fetch, analyze, and save data
def update_live_data():
    while True:
        data = fetch_top_50_cryptos()
        if data:
            df = process_data(data)
            analysis_results, top_5_by_market_cap = analyze_data(df)
            save_to_excel(df, analysis_results, top_5_by_market_cap)
            print("Data and analysis updated and saved to Excel.")
        time.sleep(300)  # Update every 5 minutes (300 seconds)

# Run the script
if __name__ == "__main__":
    update_live_data()

    